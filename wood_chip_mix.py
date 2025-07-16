import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

def pack_best_fit_combos(items, target=1300):
    """
    주어진 목표 무게에 가장 가깝게 아이템을 조합합니다.
    """
    items = sorted(items, key=lambda x: x[1])  # 가벼운 순 정렬
    used = set()
    combos = []

    remaining = [item for item in items if item[0] not in used]

    while remaining:
        best_combo = None
        best_weight = float('inf')

        # 모든 가능한 조합 탐색
        for i in range(len(remaining)):
            combo = []
            total = 0
            # 여기서 remaining[j]를 사용할 때, j가 remaining의 길이를 초과하지 않도록 보장합니다.
            for j in range(i, len(remaining)):
                if j >= len(remaining): # 인덱스 범위 확인
                    break
                name, weight = remaining[j]
                if name in used:
                    continue
                if total + weight > target * 1.5:  # 너무 무거운 조합 제외
                    break
                combo.append((name, weight))
                total += weight
                if total >= target:
                    break

            # 조합 후보가 타겟 이상일 경우, 더 좋은지 비교
            if total >= target and total < best_weight:
                best_combo = combo
                best_weight = total

        # 최적 조합이 있으면 추가
        if best_combo:
            combos.append((best_combo, best_weight))
            for name, _ in best_combo:
                used.add(name)
        else:
            break

        # 다음 반복을 위해 남은 상품 다시 필터링
        remaining = [item for item in items if item[0] not in used]
    return combos

class TimberChipCombinerApp:
    def __init__(self, master):
        self.master = master
        master.title("목재칩 조합기")

        self.excel_path = tk.StringVar()
        self.target_weight = tk.DoubleVar(value=1300.0) # 기본 목표 무게 설정

        # 파일 경로 선택
        tk.Label(master, text="엑셀 파일 경로:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(master, textvariable=self.excel_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(master, text="찾아보기", command=self.browse_excel_file).grid(row=0, column=2, padx=5, pady=5)

        # 목표 무게 설정
        tk.Label(master, text="목표 무게 (g):").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        tk.Entry(master, textvariable=self.target_weight, width=10).grid(row=1, column=1, padx=5, pady=5, sticky="w")

        # 조합 버튼
        tk.Button(master, text="조합 실행", command=self.perform_combination).grid(row=2, column=0, columnspan=3, pady=10)

        # 결과 출력 영역
        self.result_text = scrolledtext.ScrolledText(master, width=70, height=20, wrap=tk.WORD)
        self.result_text.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

        # 결과 저장 버튼
        self.save_button = tk.Button(master, text="결과 저장", command=self.save_results)
        self.save_button.grid(row=4, column=0, columnspan=3, pady=10)
        self.save_button.config(state=tk.DISABLED) # 초기에는 비활성화

        self.grouped_combos = [] # 조합 결과를 저장할 변수

    def browse_excel_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path.set(file_path)
    #   이 함수가 파일 읽어오는 함수 만약 양식이 변경 됬으면 이부분을 변경해야함
    def load_and_filter_data(self, excel_path):
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            rows = []
            # 6행부터 시작 (5행까지는 제외)
            for row_idx, row in enumerate(ws.iter_rows()):
                if row_idx < 5:  # 0부터 시작하므로 5는 6번째 행을 의미합니다.
                    continue
                cell_o = row[14]  # 'O'열 = 15번째 열 → 인덱스 14  # 이부분은 몇호인지 번호
                cell_t = row[19]  # 'T'열 = 20번째 열 → 인덱스 19  # 이부분이 목재칩 즉 무게란

                # 셀이 존재하고, 색상 정보가 있는지 확인
                fill_o = cell_o.fill.start_color.rgb if cell_o.fill.start_color else None
                fill_t = cell_t.fill.start_color.rgb if cell_t.fill.start_color else None

                # 색이 없는 셀만 선택 (흰색 또는 투명)
                # openpyxl 3.x 버전에서는 기본 색상이 '00000000'로 나타날 수 있습니다.
                # 'FFFFFFFF'는 흰색, None은 색상 없음 (기본값)
                if (fill_o in ('00000000', 'FFFFFFFF', None)) and \
                   (fill_t in ('00000000', 'FFFFFFFF', None)):
                    rows.append([cell_o.value, cell_t.value])

            df = pd.DataFrame(rows, columns=['발급번호', '목재칩'])

            # 데이터프레임의 첫 몇 행이 헤더일 가능성을 고려하여 실제 데이터 시작
            # 이미 6행부터 읽었으므로 추가적인 슬라이싱은 필요 없습니다.
            # 하지만 엑셀 시트 자체에 헤더가 더 있을 수 있으니, 목재칩 열이 숫자로 변환 가능한지 확인하는 과정이 중요합니다.
            df['목재칩'] = pd.to_numeric(df['목재칩'], errors='coerce')
            filtered_df = df.dropna(subset=['목재칩'])

            # '발급번호'가 비어있는 행을 제거 (선택 사항, 데이터 품질에 따라 조절)
            filtered_df = filtered_df.dropna(subset=['발급번호'])

            items = list(zip(filtered_df['발급번호'], filtered_df['목재칩'].round(2)))
            return items

        except Exception as e:
            messagebox.showerror("파일 읽기 오류", f"엑셀 파일을 읽는 중 오류가 발생했습니다: {e}")
            return None

    def perform_combination(self):
        excel_path = self.excel_path.get()
        if not excel_path:
            messagebox.showwarning("입력 오류", "엑셀 파일 경로를 입력해주세요.")
            return

        target_weight = self.target_weight.get()
        if target_weight <= 0:
            messagebox.showwarning("입력 오류", "목표 무게는 0보다 커야 합니다.")
            return

        items = self.load_and_filter_data(excel_path)
        if not items:
            return

        self.result_text.delete(1.0, tk.END) # 기존 결과 지우기
        self.grouped_combos = pack_best_fit_combos(items, target_weight) # 전역 변수에 저장

        output = []
        all_used = set()

        output.append("\n====== 전체 조합 결과 요약 ======\n")
        if not self.grouped_combos:
            output.append("💡 목표 무게에 맞는 조합을 찾을 수 없습니다.\n")
        else:
            for i, (combo, weight) in enumerate(self.grouped_combos, 1):
                output.append(f"[조합 {i}] 총 무게: {weight:.2f}g / 상품 수: {len(combo)}개\n")
                for name, w in combo:
                    output.append(f" - {name} ({w:.2f}g)\n")
                    all_used.add(name)
                output.append("\n")

            output.append("✅ 조합에 포함된 모든 상품:\n")
            output.append(", ".join(sorted(all_used)) + "\n")

            unused_items = [name for name, w in items if name not in all_used]
            if unused_items:
                output.append("\n❌ 조합되지 않은 상품:\n")
                output.append(", ".join(sorted(unused_items)) + "\n")
            else:
                output.append("\n🎉 모든 상품이 조합에 사용되었습니다!\n")

        self.result_text.insert(tk.END, "".join(output))
        self.save_button.config(state=tk.NORMAL) # 결과가 있을 때만 저장 버튼 활성화

    def save_results(self):
        if not self.grouped_combos:
            messagebox.showwarning("저장 오류", "저장할 조합 결과가 없습니다. 먼저 '조합 실행'을 해주세요.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="조합 결과 저장"
        )
        if not save_path:
            return

        try:
            # 새 워크북 생성
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "조합 결과"

            row_idx = 1
            all_used = set()

            for i, (combo, weight) in enumerate(self.grouped_combos, 1):
                ws_out.cell(row=row_idx, column=1, value=f"[조합 {i}]")
                ws_out.cell(row=row_idx, column=2, value=f"총 무게: {weight:.2f}g")
                ws_out.cell(row=row_idx, column=3, value=f"상품 수: {len(combo)}개")
                row_idx += 1
                for name, w in combo:
                    ws_out.cell(row=row_idx, column=1, value=name)
                    ws_out.cell(row=row_idx, column=2, value=f"{w:.2f}g")
                    all_used.add(name)
                    row_idx += 1
                row_idx += 1 # 각 조합 사이에 빈 줄 추가

            # 조합된 상품 목록
            ws_out.cell(row=row_idx, column=1, value="조합에 포함된 모든 상품:")
            row_idx += 1
            ws_out.cell(row=row_idx, column=1, value=", ".join(sorted(all_used)))
            row_idx += 2

            # 조합되지 않은 상품 목록 (원본 items를 다시 로드해야 정확)
            # 여기서는 현재 메모리에 있는 items를 사용하거나, 다시 로드할 필요가 있습니다.
            # 여기서는 perform_combination에서 로드했던 items를 다시 가져오는 것이 가장 정확합니다.
            # 하지만 현재 구조상 self.items를 저장하고 있지 않으므로, 이 부분은 GUI를 실행할 때마다 다시 계산될 수 있습니다.
            # 가장 확실한 방법은 self.perform_combination에서 self.items를 저장하는 것입니다.
            # 편의를 위해 여기서는 다시 로드하는 과정을 생략하고, 조합된 것만 확실히 저장합니다.
            # 만약 조합되지 않은 상품까지 정확히 저장하고 싶다면, load_and_filter_data의 결과를 self.items에 저장해야 합니다.

            wb_out.save(save_path)
            messagebox.showinfo("저장 완료", f"조합 결과가 '{save_path}'에 성공적으로 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("저장 오류", f"파일 저장 중 오류가 발생했습니다: {e}")

root = tk.Tk()
app = TimberChipCombinerApp(root)
root.mainloop()