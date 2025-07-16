import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color
import random


def create_test_excel_file(file_name="test_data.xlsx"):
    """
    사용자 입력에 따라 지정된 개수의 테스트 케이스를 포함하는 가상 엑셀 파일을 생성합니다.
    A열(구분), O열(발급번호), T열(무게) 데이터를 포함하며, 일부 셀에 랜덤하게 색상을 적용합니다.
    """
    while True:
        try:
            num_items = int(input("생성할 테스트 케이스 개수를 입력하세요: "))
            if num_items > 0:
                break
            else:
                print("0보다 큰 숫자를 입력해주세요.")
        except ValueError:
            print("올바른 숫자를 입력해주세요.")

    data = []
    categories = ['Category_A', 'Category_B', 'Category_C', 'Category_D', 'Category_E']

    for i in range(num_items):
        category = np.random.choice(categories)
        issue_num = f"WOOD_CHIP_{i + 1:03d}"
        weight = round(np.random.uniform(50, 300), 2)
        data.append({
            'A_COL': category,
            'O_COL': issue_num,
            'T_COL': weight
        })

    df = pd.DataFrame(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 헤더 행 추가 (GUI 프로그램의 5행 스킵 조건과 맞춤)
    # 실제 데이터는 6행부터 시작합니다.
    ws.append(['구분', '', '', '', '', '', '', '', '', '', '', '', '', '', '발급번호', '', '', '', '', '', '무게'])
    for r in range(2, 6):  # 1행은 위에 추가된 헤더, 2~5행에 더미 데이터
        ws.cell(row=r, column=1, value=f"Header_{r - 1}")
        ws.cell(row=r, column=15, value=f"Header_{r - 1}")
        ws.cell(row=r, column=20, value=f"Header_{r - 1}")

    start_row_for_data = 6  # 실제 데이터가 시작될 엑셀 행 번호

    # 랜덤하게 색상을 적용할 아이템 수 (전체 아이템의 약 20%에 색상 적용)
    number_of_colored = int(num_items * 0.2)
    # 색상을 적용할 아이템의 인덱스를 무작위로 선택
    indices_to_color = random.sample(range(num_items), number_of_colored)

    for r_idx, row_data in df.iterrows():
        # 데이터프레임의 인덱스(r_idx)에 6을 더하여 실제 엑셀 행 번호를 계산
        excel_row_number = start_row_for_data + r_idx

        ws.cell(row=excel_row_number, column=1, value=row_data['A_COL'])  # A열 (구분)
        ws.cell(row=excel_row_number, column=15, value=row_data['O_COL'])  # O열 (발급번호)
        ws.cell(row=excel_row_number, column=20, value=row_data['T_COL'])  # T열 (무게)

        # 무작위로 선택된 아이템에만 색상 적용
        if r_idx in indices_to_color:
            # 랜덤한 색상 생성
            random_color = Color(
                rgb=f"{random.randint(0, 255):02X}{random.randint(0, 255):02X}{random.randint(0, 255):02X}")
            fill = PatternFill(start_color=random_color, end_color=random_color, fill_type="solid")

            ws.cell(row=excel_row_number, column=1).fill = fill  # A열에 색상 적용
            ws.cell(row=excel_row_number, column=15).fill = fill  # O열에 색상 적용
            ws.cell(row=excel_row_number, column=20).fill = fill  # T열에 색상 적용

    # 각 열의 너비 자동 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['T'].width = 15

    wb.save(file_name)
    print(f"\n'{file_name}' 파일이 성공적으로 생성되었습니다.")
    print(f"총 {num_items}개의 항목이 포함되어 있으며, {len(indices_to_color)}개의 항목에 색상 구분이 적용되었습니다.")


if __name__ == "__main__":
    create_test_excel_file()